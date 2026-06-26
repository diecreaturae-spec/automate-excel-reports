import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Стили
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_wrap = Alignment(vertical='center', wrap_text=True)
base_font = Font(name='Times New Roman', size=12)
header_font = Font(name='Times New Roman', size=12, bold=True)
column4_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

def format_sheet(ws, fill_d_column=False):
    last_row = ws.max_row
    target_col_idx = None
    for cell in ws[1]:
        if cell.value and "Работников, устроенных по основному месту работы" in str(cell.value):
            target_col_idx = cell.column
            break

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_wrap
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=ws.max_column):
        for cell in row:
            cell.font = base_font
            cell.alignment = data_wrap
            cell.border = thin_border
            if cell.column == 3 and cell.value is not None:
                cell.value = str(cell.value).upper()
            if fill_d_column and cell.column == 4:
                cell.fill = column4_fill
            if target_col_idx and cell.column == target_col_idx:
                cell.alignment = center_wrap
            if cell.row == last_row:
                cell.font = header_font

    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                length = len(str(cell.value))
            except:
                length = 10
            if length > max_length:
                max_length = length
        adjusted = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted

def build_mapping(main_file):
    """Группировка по коду МО, сохранение листа и возврат словаря"""
    df = pd.read_excel(main_file)
    grouped = df.groupby('Код МО')['Сотрудник: ФИО сотрудника'].count().reset_index(
        name='Количество по полю Сотрудник: ФИО сотрудника'
    )
    with pd.ExcelWriter(main_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        grouped.to_excel(writer, sheet_name='Группировка по организациям', index=False)
        format_sheet(writer.sheets['Группировка по организациям'])
    return grouped.set_index('Код МО').to_dict()['Количество по полю Сотрудник: ФИО сотрудника']

def process_application(input_file, output_file, mapping_dict,
                        code_col, target_col,
                        filter_col_index,
                        total_label_col=None,
                        extra_col_to_drop=None,
                        clear_code_col=False,
                        fill_d=False):

    df = pd.read_excel(input_file)

    if extra_col_to_drop and extra_col_to_drop in df.columns:
        df.drop(columns=[extra_col_to_drop], inplace=True)

    # Определяем столбец для фильтрации и для итоговой метки
    filter_col_name = df.columns[filter_col_index]
    if total_label_col is None:
        total_label_col = filter_col_name

    # Удаляем старые итоги
    df = df[~df[filter_col_name].astype(str).str.contains('ВСЕГО', case=False, na=False)]

    # Подтягиваем данные
    df[target_col] = df[code_col].map(mapping_dict).fillna(0).astype(int)

    # Добавляем итоговую строку
    total = df[target_col].sum()
    new_row = {col: None for col in df.columns}
    new_row[target_col] = total
    new_row[total_label_col] = 'ВСЕГО:'
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    # Для №3 очищаем столбец с кодом
    if clear_code_col and code_col in df.columns:
        df[code_col] = ''

    # Сохраняем
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        format_sheet(writer.sheets['Sheet1'], fill_d_column=fill_d)


if __name__ == '__main__':
    mapping = build_mapping('Бережливое_производство.xlsx')
    print(f'Справочник построен, записей: {len(mapping)}')

    # 2. Приложение №1
    process_application(
        input_file='Приложение №1.xlsx',
        output_file='Приложение №1_готово.xlsx',
        mapping_dict=mapping,
        code_col='Код МО',
        target_col='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства',
        filter_col_index=2,
        extra_col_to_drop='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства, ед.',
        fill_d=True
    )

    # 3. Приложение №2
    process_application(
        input_file='Приложение №2.xlsx',
        output_file='Приложение №2_готово.xlsx',
        mapping_dict=mapping,
        code_col='Код МО',
        target_col='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства',
        filter_col_index=2,
        extra_col_to_drop='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства, ед.',
        fill_d=True
    )

    # 4. Приложение №3
    process_application(
        input_file='Приложение №3(с кодом).xlsx',
        output_file='Приложение №3_готово.xlsx',
        mapping_dict=mapping,
        code_col='Период (месяц из выпадающего списка)',
        target_col='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства',
        filter_col_index=3,
        total_label_col='Наименование МО',
        extra_col_to_drop='Работников, устроенных по основному месту работы в МО, обученных по вопросам применения бережливого производства, в шт',
        clear_code_col=True,
        fill_d=False
    )

    print('Готово!')
